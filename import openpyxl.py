import openpyxl

# Путь к файлу Excel
excel_file_path = 'C:\\Users\\Dima\\YandexDisk\\Школа\\ЕГЭ_1\\answers.xlsx'

# Открываем файл Excel
workbook = openpyxl.load_workbook(excel_file_path)

# Выбираем лист (sheet)
sheet = workbook['Лист1']

# Номер строки и столбца для извлечения данных
row_number = 2  # Например, строка №2
column_number = 10  # Например, столбец B
def find_column(number_task):
    for i in range(1,30):
        if str(sheet.cell(1,i).value) == str(number_task):
            print(sheet.cell(1,i).value)
            break
    return i

column_number=find_column(17)

# Получаем значение из ячейки
cell_value = sheet.cell(row_number+1,column_number).value

# Выводим значение
print(f'Номер {17} Задание {row_number}: {cell_value}')

# Закрываем файл Excel
workbook.close()
