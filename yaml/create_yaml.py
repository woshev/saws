import yaml
import ruamel.yaml
from docx import Document
import re
import openpyxl
import os
import textwrap


# Создайте объект YAML
yaml = ruamel.yaml.YAML(typ='rt')
base_folder = 'D:\\База\\YandexDisk\\Школа\\ЕГЭ_1'
task_file_name = base_folder+'\\e.docx'
answer_file_name =base_folder+'\\answers.xlsx'
write_file_name=base_folder+'\\17.yaml'

def find_column(number_task):
    for i in range(1,30):
        if str(sheet.cell(1,i).value) == str(number_task):
            break
    return i


# Открываем файл Excel
workbook = openpyxl.load_workbook(answer_file_name)
# Выбираем лист (sheet)
sheet = workbook['Лист1']

# Загрузите YAML файл
with open(write_file_name, 'r', encoding='utf-8') as file:
    yaml1 = ruamel.yaml.YAML()
    yaml1.default_style=''
    yaml1.width = 20000
    data = yaml1.load(file)

# Пример использования с вашим путем к файлу
docx_file_path = task_file_name
doc = Document(docx_file_path)


answer_workbook=openpyxl.load_workbook


def filetotasks(doc):
    tmp=['']
    text=''
    last=''

    for i in doc.paragraphs:
        if i.text=='':
            break
        if  i.text[0] in ['0', '1', '2','3','4','5','6','7','8','9']:
            
            if last != '' and last[0] in ['0', '1', '2','3','4','5','6','7','8','9'] and tmp[-1] != last:
                tmp.append(last)
            if text != '' and tmp[-1] != text:
                tmp[-1] = text
                text=''
            last = i.text
            
        else:
            if text == '':
                text=last
            text+=i.text
        if last != '' and last[0] in ['0', '1', '2','3','4','5','6','7','8','9'] and tmp[-1] != last:
                tmp.append(last)
    tmp =tmp[1:]
    return tmp

def split_task_and_write(task,ege_type):
    #print('\n t\n\n',t)
    number_kpolyakov=''
    author=''
    answer=''
    input_data=''
    input_file=''
    solve_text_python=''
    solve_file_python=''
    solve_text_pascal=''
    solve_file_pascal=''
    youtube_solve=''
    while task[0] in ['0', '1', '2','3','4','5','6','7','8','9']:
        number_kpolyakov+=task[0]
        task=task[1:]
    number_kpolyakov=int(number_kpolyakov)
    while task[0] in ['\t',')']:
        task=task[1:]
    if task[0] == '(':
        task=task[1:]
        while task[0] != ')':
            author+=task[0]
            task=task[1:]
        while task[0] in ['\t',')',' ']:
            task=task[1:]

    answer=sheet.cell(number_kpolyakov+1,find_column(ege_type)).value
    print(answer)
    if type(answer) == str:
        answer = [line.strip().replace('_x000D_', '') for line in answer.splitlines()]
    elif type(answer) == int:
        answer = str(answer)
    else:
        print("type answer error")
        print("answer:\n",answer)
    if len(answer) > 1:
        youtube_solve=answer[-1]
    task = '\n'.join(textwrap.wrap(task, width=80, break_long_words=False))
    print(task)
    answer=answer[0]
    #поиск файла с исходными данными
    input_file = re.findall(r'\b\d+-\d+\.txt\b',task)
    
    if input_file != []:
        input_file=input_file[0]
        f = open(base_folder+'\\17data\\'+str(input_file), 'r',encoding='utf-8')
        input_data =  f.read()
        f.close()
        #print(input_data)
    #импорт текст программы Pascal
    if os.path.exists(base_folder+'\\'+str(ege_type)+'solve\\'+str(ege_type)+'-'+str(number_kpolyakov)+'.pas'):
        solve_file_pascal=str(number_kpolyakov)+'.pas'
        f_pas=open(base_folder+'\\'+str(ege_type)+'solve\\'+str(ege_type)+'-'+str(number_kpolyakov)+'.pas','r',encoding='utf-8')
        solve_text_pascal=f_pas.read()
        f_pas.close()
        #импорт текст программы Python
    if os.path.exists(base_folder+'\\'+str(ege_type)+'solve\\'+str(ege_type)+'-'+str(number_kpolyakov)+'.py'):
        solve_file_python=str(number_kpolyakov)+'.py'
        f_pas=open(base_folder+'\\'+str(ege_type)+'solve\\'+str(ege_type)+'-'+str(number_kpolyakov)+'.py','r',encoding='utf-8')
        solve_text_python=f_pas.read()
        f_pas.close()
    #print(solve_text_pascal)
    #print(solve_file_pascal)
    #print(solve_text_python)
    #print(solve_file_python)
    #print("номер\n",number_kpolyakov)
    #print(file)
    #print(task)
    #print(answer)
    #print(youtube_solve)
    #print("имя\n",author)
    #print("Задание\n",task)
    topic=None
    task_type_id=None
    return {'task id':str(ege_type)+'_'+str(number_kpolyakov),
         'ege type':ege_type,
         'topic':None if topic==None else topic,
         'task type id':None if task_type_id==None else task_type_id,
         'author': None if author=='' else author,
         'number kpolyakov':number_kpolyakov,
         'task description': task,
         'input data':None if input_data == '' else input_data,
         'input file':None if input_file == [] else input_file,
         'answer':None if answer == '' else answer,
         'solve text Python':None if solve_text_python == '' else solve_text_python,
         'solve file Python':None if solve_file_python == '' else solve_file_python,
         'solve text Pascal':None if solve_text_pascal == '' else solve_text_pascal,
         'solve file Pascal':None if solve_file_pascal == '' else solve_file_pascal,                               
         'solve youtube':None if youtube_solve == '' else youtube_solve
    }




def writetoyaml(data,data_task):
    try:
        l = len(data[data_task['ege type']])
    except:
        l=1            
        data[data_task['ege type']] = [{'task id':'',
                'ege type':'',
                'topic':'',
                'task type id':'',
                'author':'',
                'number kpolyakov':'',
                'task description': '',
                'input data':'',
                'input file':'',
                'answer':'',
                'solve text Python':'',
                'solve file Python':'',
                'solve text Pascal':'',
                'solve file Pascal':'',                               
                'solve youtube':''
                }]
    if data_task['number kpolyakov']-1 < l:
        
        # Измените значение "условие" с использованием SingleQuotedScalarString
        data[data_task['ege type']][data_task['number kpolyakov']-1]['task id'] = data_task['task id']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['ege type'] = data_task['ege type']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['topic'] = data_task['topic']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['task type id'] = data_task['task type id']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['author'] = data_task['author']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['number kpolyakov'] = data_task['number kpolyakov']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['task description'] = ruamel.yaml.scalarstring.LiteralScalarString(data_task['task description'])
        data[data_task['ege type']][data_task['number kpolyakov']-1]['input data'] = data_task['input data']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['input file'] = data_task['input file']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['answer'] = data_task['answer']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['solve text Python'] = ruamel.yaml.scalarstring.LiteralScalarString(data_task['solve text Python'])
        data[data_task['ege type']][data_task['number kpolyakov']-1]['solve file Python'] = data_task['solve file Python']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['solve text Pascal'] = ruamel.yaml.scalarstring.LiteralScalarString(data_task['solve text Pascal'])
        data[data_task['ege type']][data_task['number kpolyakov']-1]['solve file Pascal'] = data_task['solve file Pascal']
        data[data_task['ege type']][data_task['number kpolyakov']-1]['solve youtube'] = data_task['solve youtube']
        
    else:
        data[data_task['ege type']].append({'task id':data_task['task id'],
                                        'ege type':data_task['ege type'],
                                        'topic':data_task['topic'],
                                        'task type id':data_task['task type id'],
                                        'author':data_task['author'],
                                        'number kpolyakov':data_task['number kpolyakov'],
                                        'task description': ruamel.yaml.scalarstring.LiteralScalarString(data_task['task description']),
                                        'input data':data_task['input data'],
                                        'input file':data_task['input file'],
                                        'answer':data_task['answer'],
                                        'solve text Python':ruamel.yaml.scalarstring.LiteralScalarString(data_task['solve text Python']),
                                        'solve file Python':data_task['solve file Python'],
                                        'solve text Pascal':ruamel.yaml.scalarstring.LiteralScalarString(data_task['solve text Pascal']),
                                        'solve file Pascal':data_task['solve file Pascal'],                               
                                        'solve youtube':data_task['solve youtube']
                                            })
                        



for i in filetotasks(doc):
    l = split_task_and_write(i,17)
    writetoyaml(data,l),
#split_task_and_write(filetotasks(doc),17)
with open(write_file_name, 'w', encoding='utf-8') as file:

    yaml1.dump(data, file)

workbook.close()






'''import ruamel.yaml

# Создайте объект YAML
yaml = ruamel.yaml.YAML(typ='rt')

# Загрузите YAML файл
with open('d:/База/YandexDisk/Школа/ЕГЭ_1/1.yaml', 'r', encoding='utf-8') as file:
    data = yaml.load(file)

# Измените значение "условие" с использованием SingleQuotedScalarString
data['Номер 17'][0]['условие'] = ruamel.yaml.scalarstring.FoldedScalarString('Новое значение условия')

# Сохраните обновленный YAML файл с разрешением использования Unicode, оригинальным порядком ключей
# и с использованием SingleQuotedScalarString для поля "условие"
with open('d:/База/YandexDisk/Школа/ЕГЭ_1/1.yaml', 'w', encoding='utf-8') as file:
    yaml.dump(data, file)
'''


'''print(d)
tmp=[]
for i in d:
    for j in d[i]:
        print(j["номер поляков"]+'\n')
        print(j["подтип"]+'\n')
        if j["подтип"] not in tmp:
            tmp.append(j["подтип"])
        #for k in j["тип"]:
        #    print(k)
            
        #print('\n'+j["условие"]+'\n\n')
print(tmp)'''
